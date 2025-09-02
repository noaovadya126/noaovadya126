import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os

class KoreanWordProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Korean Word Processor - Simple Version")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')
        
        self.file_path = ""
        self.workbook = None
        self.worksheet = None
        
        self.setup_ui()
        
    def setup_ui(self):
        title_label = tk.Label(self.root, text="Korean Word Processor - Simple Version", font=("Arial", 16, "bold"), bg='#f0f0f0')
        title_label.pack(pady=20)
        
        file_frame = tk.Frame(self.root, bg='#f0f0f0')
        file_frame.pack(pady=10, padx=20, fill='x')
        
        tk.Label(file_frame, text="Excel File:", font=("Arial", 10), bg='#f0f0f0').pack(anchor='w')
        
        file_select_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file_select_frame.pack(fill='x', pady=5)
        
        self.file_entry = tk.Entry(file_select_frame, font=("Arial", 10), width=60)
        self.file_entry.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        browse_btn = tk.Button(file_select_frame, text="Browse", command=self.browse_file, 
                              bg='#4CAF50', fg='white', font=("Arial", 10), relief='flat')
        browse_btn.pack(side='right')
        
        button_frame = tk.Frame(self.root, bg='#f0f0f0')
        button_frame.pack(pady=20)
        
        self.process_btn = tk.Button(button_frame, text="Process All", command=self.process_file, 
                                    bg='#2196F3', fg='white', font=("Arial", 12, "bold"), 
                                    relief='flat', width=15, height=2)
        self.process_btn.pack(side='left', padx=10)
        
        self.test_btn = tk.Button(button_frame, text="Test First 50", command=self.test_first_50, 
                                 bg='#9C27B0', fg='white', font=("Arial", 12, "bold"), 
                                 relief='flat', width=15, height=2)
        self.test_btn.pack(side='left', padx=10)
        
        self.save_btn = tk.Button(button_frame, text="Save As", command=self.save_file, 
                                 bg='#FF9800', fg='white', font=("Arial", 12, "bold"), 
                                 relief='flat', width=15, height=2, state='disabled')
        self.save_btn.pack(side='left', padx=10)
        
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(pady=20, padx=20, fill='x')
        
        self.status_label = tk.Label(self.root, text="Ready", font=("Arial", 10), bg='#f0f0f0')
        self.status_label.pack(pady=10)
        
        preview_frame = tk.Frame(self.root, bg='#f0f0f0')
        preview_frame.pack(pady=10, padx=20, fill='both', expand=True)
        
        tk.Label(preview_frame, text="Preview:", font=("Arial", 10, "bold"), bg='#f0f0f0').pack(anchor='w')
        
        self.preview_text = tk.Text(preview_frame, height=12, width=90, font=("Consolas", 9))
        scrollbar = tk.Scrollbar(preview_frame, orient="vertical", command=self.preview_text.yview)
        self.preview_text.configure(yscrollcommand=scrollbar.set)
        
        self.preview_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path = file_path
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.status_label.config(text="File selected: " + os.path.basename(file_path))
            
    def test_first_50(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please select a file first!")
            return
            
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            max_row = self.worksheet.max_row
            test_rows = min(50, max_row - 1)
            
            self.progress['maximum'] = test_rows
            self.progress['value'] = 0
            
            self.status_label.config(text=f"Testing first {test_rows} words... Please wait.")
            self.root.update()
            
            preview_data = []
            
            for row_num in range(2, test_rows + 2):
                korean_word = str(self.worksheet.cell(row=row_num, column=4).value or "")
                pos = str(self.worksheet.cell(row=row_num, column=5).value or "")
                original_guide = str(self.worksheet.cell(row=row_num, column=6).value or "")
                
                ai_output = self.get_ai_assistance(korean_word, pos)
                
                preview_data.append(f"{row_num-1:2d}. {korean_word:12s} ({pos:6s}) | Original: {original_guide:25s} | AI: {ai_output}")
                
                self.progress['value'] = row_num - 1
                self.root.update()
                
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, "\n".join(preview_data))
            
            self.status_label.config(text=f"Test completed! Processed {test_rows} words.")
            self.save_btn.config(state='normal')
            messagebox.showinfo("Success", f"Test completed successfully! Processed {test_rows} words.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error processing file: {str(e)}")
            self.status_label.config(text="Error occurred")
            
    def process_file(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please select a file first!")
            return
            
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            max_row = self.worksheet.max_row
            self.progress['maximum'] = max_row - 1
            self.progress['value'] = 0
            
            self.status_label.config(text="Processing... Please wait.")
            self.root.update()
            
            for row_num in range(2, max_row + 1):
                korean_word = str(self.worksheet.cell(row=row_num, column=4).value or "")
                pos = str(self.worksheet.cell(row=row_num, column=5).value or "")
                
                ai_output = self.get_ai_assistance(korean_word, pos)
                
                self.worksheet.cell(row=row_num, column=9, value=ai_output)
                
                self.progress['value'] = row_num - 1
                if row_num % 100 == 0:
                    self.root.update()
                
            self.status_label.config(text="Processing completed!")
            self.save_btn.config(state='normal')
            messagebox.showinfo("Success", "File processed successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error processing file: {str(e)}")
            self.status_label.config(text="Error occurred")
            
    def get_ai_assistance(self, korean_word, pos):
        try:
            if pos == "명사":
                return f"{korean_word}을 사용하다"
            elif pos == "동사":
                return f"{korean_word}을 하다"
            elif pos == "형용사":
                return f"{korean_word} 상태이다"
            elif pos == "부사":
                return f"{korean_word} 행동하다"
            else:
                return f"{korean_word}의 예시"
        except Exception as e:
            return f"Error: {str(e)}"
            
    def save_file(self):
        if self.workbook is None:
            messagebox.showerror("Error", "No processed data to save!")
            return
            
        save_path = filedialog.asksaveasfilename(
            title="Save Processed File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if save_path:
            try:
                self.workbook.save(save_path)
                messagebox.showinfo("Success", f"File saved successfully to:\n{save_path}")
                self.status_label.config(text="File saved successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Error saving file: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = KoreanWordProcessor(root)
    root.mainloop()
